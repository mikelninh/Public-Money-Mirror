"""Export services for different data formats."""
import io
from typing import Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_to_csv(data: Dict[str, Any], drilldown: str = "function") -> bytes:
    """
    Export spending data to CSV format.
    
    Args:
        data: Spending data dictionary with distribution and percentages
        drilldown: Dimension name for context
    
    Returns:
        Bytes of CSV content
    """
    dist = data.get("distribution", {})
    pcts = data.get("percentages", {})
    
    # Create DataFrame
    df = pd.DataFrame([
        {
            "Kategorie": category,
            "Betrag (EUR)": amount,
            "Prozent": pcts.get(category, 0),
            "Betrag (Milliarden EUR)": round(amount / 1e9, 2)
        }
        for category, amount in sorted(dist.items(), key=lambda x: x[1], reverse=True)
    ])
    
    # Convert to CSV
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False, encoding='utf-8')
    return csv_buffer.getvalue().encode('utf-8')


def export_to_excel(data: Dict[str, Any], drilldown: str = "function") -> bytes:
    """
    Export spending data to Excel format with formatting.
    
    Args:
        data: Spending data dictionary with distribution and percentages
        drilldown: Dimension name for context
    
    Returns:
        Bytes of Excel content
    """
    dist = data.get("distribution", {})
    pcts = data.get("percentages", {})
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bundeshaushalt 2024"
    
    # Header styling
    header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Write headers
    headers = ["Kategorie", "Betrag (EUR)", "Prozent", "Betrag (Milliarden EUR)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    sorted_items = sorted(dist.items(), key=lambda x: x[1], reverse=True)
    for row_num, (category, amount) in enumerate(sorted_items, 2):
        ws.cell(row=row_num, column=1, value=category)
        ws.cell(row=row_num, column=2, value=amount)
        ws.cell(row=row_num, column=3, value=round(pcts.get(category, 0), 2))
        ws.cell(row=row_num, column=4, value=round(amount / 1e9, 2))
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=len(sorted_items) + 1, min_col=col_num, max_col=col_num):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 50)
    
    # Add metadata sheet
    ws_meta = wb.create_sheet("Metadaten")
    ws_meta.append(["Quelle", data.get("source", "Intern")])
    ws_meta.append(["Datenqualität", data.get("data_quality", "Nicht verfügbar")])
    ws_meta.append(["Letztes Update", data.get("last_updated", "2024")])
    if "total" in data:
        ws_meta.append(["Gesamtbudget", f"{data['total']:,.0f} EUR"])
    
    # Save to bytes
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    return excel_buffer.getvalue()


def export_bundeslaender_to_csv(data: Dict[str, Any]) -> bytes:
    """
    Export Bundesländer data to CSV.
    
    Args:
        data: Bundesländer data dictionary
    
    Returns:
        Bytes of CSV content
    """
    states = data.get("states", {})
    
    rows = []
    for state_name, state_data in states.items():
        rows.append({
            "Bundesland": state_name,
            "Budget (EUR)": state_data.get("budget_allocation", 0),
            "Budget (Milliarden EUR)": round(state_data.get("budget_allocation", 0) / 1e9, 2),
            "Bevölkerung": state_data.get("population", 0),
            "Pro Kopf": round(state_data.get("budget_allocation", 0) / max(state_data.get("population", 1), 1), 2),
            "Hauptkategorie": state_data.get("top_category", "N/A")
        })
    
    df = pd.DataFrame(rows)
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False, encoding='utf-8')
    return csv_buffer.getvalue().encode('utf-8')


def export_historical_trends_to_csv(data: Dict[str, Any]) -> bytes:
    """
    Export historical trends data to CSV.
    
    Args:
        data: Historical trends dictionary
    
    Returns:
        Bytes of CSV content
    """
    years = data.get("years", [])
    trends_data = data.get("data", {})
    
    # Get all categories
    all_categories = set()
    for year_data in trends_data.values():
        all_categories.update(year_data.keys())
    
    # Create pivot table
    rows = []
    for year in sorted(years):
        row = {"Jahr": year}
        for category in sorted(all_categories):
            row[category] = trends_data.get(year, {}).get(category, 0)
        rows.append(row)
    
    df = pd.DataFrame(rows)
    csv_buffer = io.StringIO()
    df.to_csv(csv_buffer, index=False, encoding='utf-8')
    return csv_buffer.getvalue().encode('utf-8')

